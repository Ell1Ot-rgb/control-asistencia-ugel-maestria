from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)

def get_auth_token():
    res = client.post(
        "/api/v1/auth/sessions",
        json={"username": "director.demo", "password": "Demo12345"},
    )
    return res.json()["token"]

def test_parse_dat_attlog():
    token = get_auth_token()
    dat_content = (
        "45678912\t2026-08-03 07:50:00\t1\n"
        "45678912\t2026-08-03 13:00:00\t0\n"
    )
    res = client.post(
        "/api/v1/biometric-imports",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": ("attlog.dat", dat_content.encode("utf-8"), "text/plain")},
    )
    assert res.status_code == 201
    data = res.json()
    assert data["status"] == "draft"
    assert data["total_rows"] == 2

def test_export_official_excel():
    token = get_auth_token()
    res = client.get(
        "/api/v1/reports/official-excel?month=8&year=2026",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(res.content) > 1000


def test_editable_anexo03_drives_template_workbook():
    from io import BytesIO
    from openpyxl import load_workbook

    token = get_auth_token()
    headers = {"Authorization": f"Bearer {token}"}
    edit = client.patch(
        "/api/v1/reports/annex-03/attendance",
        headers=headers,
        json={
            "month": 12,
            "year": 2026,
            "staff_member_id": 1,
            "attendance_date": "2026-12-15",
            "status": "late",
            "late_minutes": 33,
        },
    )
    assert edit.status_code == 200
    assert edit.json()["late_minutes"] == 33

    report = client.get(
        "/api/v1/reports/official-excel?month=12&year=2026",
        headers=headers,
    )
    assert report.status_code == 200
    workbook = load_workbook(BytesIO(report.content), data_only=False)
    assert workbook.sheetnames[:2] == ["ASISTENCIA", "REPORTE CONSOLIDADO"]
    assert len(workbook["ASISTENCIA"].merged_cells.ranges) > 0
    assert workbook["ASISTENCIA"]["G5"].value == "DICIEMBRE"
    assert workbook["ASISTENCIA"]["D14"].value == "45678912"
    assert workbook["ASISTENCIA"]["W14"].value == 33
