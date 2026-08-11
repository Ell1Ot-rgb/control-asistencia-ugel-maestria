"""TEC-D09 + TEC-D12 — annex reports using institution header + attendance_day."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from app.services.attendance_service import VALID_ATTENDANCE_STATUSES, attendance_service
from app.services.staff_member_service import (
    StaffMemberNotFoundError,
    staff_member_service,
)

DEMO_INSTITUTION = {
    "ugel": "UGEL Demo",
    "school_name": "IE Demo CHIQUISTRUKIS",
    "modular_code": "1234567",
    "education_level": "Secundaria",
    "shift_name": "Mañana",
}


class ReportService:
    def __init__(self) -> None:
        self._attendance_overrides: dict[tuple[int, str, int], dict[str, Any]] = {}

    def reset(self) -> None:
        self._attendance_overrides.clear()

    def annex_03(
        self, month: int, year: int, institution: dict[str, Any] | None = None
    ) -> dict[str, Any]:
        inst = institution or DEMO_INSTITUTION
        attendance_rows = self._effective_attendance(month, year)
        rows_by_staff: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in attendance_rows:
            rows_by_staff[row["staff_member_id"]].append(row)

        rows = []
        active_staff = staff_member_service.list(is_active="Y")
        for staff_member in active_staff:
            staff_member_id = staff_member["id"]
            days = rows_by_staff.get(staff_member_id, [])
            rows.append(
                {
                    "staff_member_id": staff_member_id,
                    "dni": staff_member.get("dni"),
                    "full_name": self._full_name(staff_member),
                    "days": sorted(days, key=lambda day: day["attendance_date"]),
                }
            )

        return {
            "institution": inst,
            "period": {"month": month, "year": year},
            "source": "attendance_day",
            "rows": rows,
        }

    def annex_04(self, month: int, year: int) -> dict[str, Any]:
        active_staff = staff_member_service.list(is_active="Y")
        attendance_rows = self._effective_attendance(month, year)
        totals = Counter(row["status"] for row in attendance_rows)

        rows_by_staff: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in attendance_rows:
            rows_by_staff[row["staff_member_id"]].append(row)

        detail_rows = []
        for staff_member in active_staff:
            staff_member_id = staff_member["id"]
            days = rows_by_staff.get(staff_member_id, [])
            summary = Counter(day["status"] for day in days)
            detail_rows.append(
                {
                    "staff_member_id": staff_member_id,
                    "dni": staff_member.get("dni"),
                    "full_name": self._full_name(staff_member),
                    "job_title": staff_member.get("job_title"),
                    "employment_status": staff_member.get("employment_status"),
                    "summary": {
                        "present": summary["present"],
                        "late": summary["late"],
                        "absent": summary["absent"],
                        "justified": summary["justified"],
                        "leave": summary["leave"],
                        "permission": summary["permission"],
                    },
                }
            )

        return {
            "institution": DEMO_INSTITUTION,
            "period": {"month": month, "year": year},
            "source": "attendance_day",
            "staff_count": len(active_staff),
            "totals": {
                "present": totals["present"],
                "late": totals["late"],
                "absent": totals["absent"],
                "justified": totals["justified"],
                "leave": totals["leave"],
                "permission": totals["permission"],
            },
            "rows": detail_rows,
        }

    def _staff_member(self, staff_member_id: int) -> dict[str, Any]:
        try:
            return staff_member_service.get(staff_member_id)
        except StaffMemberNotFoundError:
            return {
                "dni": None,
                "last_names": "Unknown",
                "first_names": "Staff",
            }

    def _full_name(self, staff_member: dict[str, Any]) -> str:
        return f"{staff_member['last_names']}, {staff_member['first_names']}"

    def set_attendance_override(
        self,
        *,
        month: int,
        year: int,
        staff_member_id: int,
        attendance_date: str,
        status: str,
        late_minutes: int = 0,
        observation: str | None = None,
    ) -> dict[str, Any]:
        parsed_date = date.fromisoformat(attendance_date)
        if parsed_date.month != month or parsed_date.year != year:
            raise ValueError("date_outside_period")
        key = (staff_member_id, parsed_date.isoformat(), month)
        if status == "none":
            self._attendance_overrides.pop(key, None)
            return {"staff_member_id": staff_member_id, "attendance_date": parsed_date.isoformat(), "status": "none", "late_minutes": 0}
        if status not in VALID_ATTENDANCE_STATUSES:
            raise ValueError("invalid_status")
        if late_minutes < 0:
            raise ValueError("invalid_late_minutes")
        override = {
            "staff_member_id": staff_member_id,
            "attendance_date": parsed_date.isoformat(),
            "status": status,
            "late_minutes": late_minutes,
            "justification_id": None,
            "observation": observation,
        }
        self._attendance_overrides[(staff_member_id, parsed_date.isoformat(), month)] = override
        return dict(override)

    def _effective_attendance(self, month: int, year: int) -> list[dict[str, Any]]:
        rows_by_key = {
            (row["staff_member_id"], row["attendance_date"]): dict(row)
            for row in attendance_service.list_month(month, year)
        }
        for (staff_member_id, attendance_date, override_month), override in self._attendance_overrides.items():
            if override_month != month or not attendance_date.startswith(f"{year:04d}-{month:02d}-"):
                continue
            key = (staff_member_id, attendance_date)
            row = dict(override)
            row["id"] = rows_by_key.get(key, {}).get("id", 0)
            rows_by_key[key] = row
        return sorted(rows_by_key.values(), key=lambda row: (row["attendance_date"], row["staff_member_id"]))

    def export_official_excel(self, month: int, year: int) -> bytes:
        import calendar
        import io
        from datetime import date

        import openpyxl
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Font, PatternFill

        template_path = Path(__file__).resolve().parents[2] / "templates" / "PLANTILLA-INFORME-ASIST-INICIAL-2021.xlsx"
        wb = openpyxl.load_workbook(template_path)
        ws1 = wb["ASISTENCIA"]
        ws2 = wb["REPORTE CONSOLIDADO"]
        month_names = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        month_name = month_names[month]
        institution = DEMO_INSTITUTION

        # Metadata shared by both official sheets.
        for ws in (ws1, ws2):
            ws["G5"] = month_name
            ws["S5"] = year
            ws["F6"] = institution["school_name"]
            ws["F7"] = institution["education_level"]
            ws["L7"] = "Dirección de IE"
            ws["F8"] = institution["modular_code"]
            ws["K8"] = "PUNO"
            ws["Q8"] = "SAN ROMÁN"
            ws["AC5"] = institution["shift_name"]

        active_staff = staff_member_service.list(is_active="Y")[:35]
        effective_rows = self._effective_attendance(month, year)
        by_staff: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in effective_rows:
            by_staff[row["staff_member_id"]].append(row)

        # Preserve the template's styles and validations while replacing its data rows.
        for ws, end_col in ((ws1, 49), (ws2, 39)):
            for row_num in range(13 if ws is ws2 else 14, 49):
                for col_num in range(1, end_col + 1):
                    cell = ws.cell(row_num, col_num)
                    if not isinstance(cell, MergedCell):
                        cell.value = None

        status_colors = {
            "A": ("DCFCE7", "15803D"),
            "J": ("DBEAFE", "1D4ED8"),
            "I": ("FEE2E2", "B91C1C"),
            "LG": ("DBEAFE", "1D4ED8"),
            "LS": ("DBEAFE", "1D4ED8"),
            "P": ("DBEAFE", "1D4ED8"),
            "H": ("FEE2E2", "B91C1C"),
        }

        for index, staff in enumerate(active_staff, start=1):
            staff_id = staff["id"]
            rows = by_staff.get(staff_id, [])
            days = {date.fromisoformat(row["attendance_date"]).day: row for row in rows}
            excel_row = 13 + index
            ws1.cell(excel_row, 1, index)
            ws1.cell(excel_row, 4, staff.get("dni", ""))
            ws1.cell(excel_row, 5, f'{staff.get("last_names", "")}, {staff.get("first_names", "")}')
            ws1.cell(excel_row, 6, staff.get("job_title") or "Docente")
            ws1.cell(excel_row, 7, staff.get("employment_status") or "Nombrado")
            ws1.cell(excel_row, 8, "30 hrs")

            late_minutes = 0
            justified = 0
            absent = 0
            leave_without_pay = 0
            for day in range(1, 32):
                cell = ws1.cell(excel_row, 8 + day)
                row = days.get(day)
                if day > calendar.monthrange(year, month)[1] or not row:
                    value = None
                elif row["status"] == "late":
                    value = row.get("late_minutes", 0) or 15
                    late_minutes += value
                elif row["status"] == "present":
                    value = "A"
                elif row["status"] in {"justified", "permission"}:
                    value = "J"
                    justified += 1
                elif row["status"] == "leave":
                    value = "LS"
                    leave_without_pay += 1
                elif row["status"] == "absent":
                    value = "I"
                    absent += 1
                else:
                    value = None
                cell.value = value
                if value in status_colors:
                    fill, font = status_colors[value]
                    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
                    cell.font = Font(color=font, bold=True)

            ws1.cell(excel_row, 40, absent + leave_without_pay)
            ws1.cell(excel_row, 41, "Conforme RSG 326")
            # The template summary columns are intentionally populated with values so
            # the workbook is useful even before a spreadsheet engine recalculates formulas.
            ws1.cell(excel_row, 43, justified)
            ws1.cell(excel_row, 44, 0)
            ws1.cell(excel_row, 45, leave_without_pay)
            ws1.cell(excel_row, 46, absent)
            ws1.cell(excel_row, 47, 0)
            ws1.cell(excel_row, 48, sum(1 for row in rows if row["status"] == "late"))
            ws1.cell(excel_row, 49, late_minutes)

            report_row = 12 + index
            summary = Counter(row["status"] for row in rows)
            ws2.cell(report_row, 1, index)
            ws2.cell(report_row, 4, staff.get("dni", ""))
            ws2.cell(report_row, 5, f'{staff.get("last_names", "")}, {staff.get("first_names", "")}')
            ws2.cell(report_row, 6, staff.get("job_title") or "Docente")
            ws2.cell(report_row, 7, staff.get("employment_status") or "Nombrado")
            ws2.cell(report_row, 8, "30 hrs")
            ws2.cell(report_row, 10, summary["justified"])
            ws2.cell(report_row, 12, 0)
            ws2.cell(report_row, 13, summary["leave"])
            ws2.cell(report_row, 16, summary["absent"])
            ws2.cell(report_row, 18, late_minutes)
            ws2.cell(report_row, 20, late_minutes // 60)
            ws2.cell(report_row, 21, late_minutes % 60)
            ws2.cell(report_row, 23, 0)
            ws2.cell(report_row, 34, summary["absent"] + summary["leave"])
            ws2.cell(report_row, 35, "Conforme RSG 326")

        # Keep the workbook linked and force Excel/Sheets to recalculate formulas on open.
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

report_service = ReportService()
